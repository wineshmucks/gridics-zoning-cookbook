from __future__ import annotations

import re
import json
from dataclasses import dataclass
from typing import Protocol
from pathlib import Path
from urllib.parse import parse_qs, urljoin, urlparse

import httpx
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from zoning_agno.models.schemas import (
    Citation,
    DefinedTerm,
    SourceChunk,
    SourceDocument,
    SourceKind,
    SourceSection,
    SourceTable,
)


@dataclass(frozen=True)
class _ParsedTable:
    title: str
    headers: list[str]
    rows: list[list[str]]
    footnotes: list[str]


class SourceResolver(Protocol):
    def can_resolve(self, source_url: str, source_kind: SourceKind) -> bool: ...

    def resolve(self, source_url: str, jurisdiction: str, source_kind: SourceKind) -> SourceDocument: ...


class GenericHtmlResolver:
    def can_resolve(self, source_url: str, source_kind: SourceKind) -> bool:
        if source_kind == SourceKind.EXCEL:
            return False
        return source_kind in {SourceKind.MUNICODE, SourceKind.HTML} or source_url.startswith("http")

    def resolve(self, source_url: str, jurisdiction: str, source_kind: SourceKind) -> SourceDocument:
        return _load_html_source(source_url, jurisdiction, source_kind)


class ResolverRegistry:
    def __init__(self, resolvers: list[SourceResolver] | None = None) -> None:
        self.resolvers = resolvers or [ExcelResolver(), GenericHtmlResolver()]

    def resolve(self, source_url: str, jurisdiction: str, source_kind: SourceKind) -> SourceDocument:
        for resolver in self.resolvers:
            if resolver.can_resolve(source_url, source_kind):
                return resolver.resolve(source_url, jurisdiction, source_kind)
        return SourceDocument(
            jurisdiction=jurisdiction,
            source_kind=source_kind,
            source_url=source_url,
            document_title=f"No resolver available for {source_kind.value}",
            sections=[],
            tables=[],
            definitions=[],
            metadata={"parser": "unresolved"},
        )


def load_source_document(source_url: str, jurisdiction: str, source_kind: SourceKind) -> SourceDocument:
    return ResolverRegistry().resolve(source_url, jurisdiction, source_kind)


class ExcelResolver:
    def can_resolve(self, source_url: str, source_kind: SourceKind) -> bool:
        return source_kind == SourceKind.EXCEL or Path(source_url).suffix.lower() in {".xlsx", ".xlsm"}

    def resolve(self, source_url: str, jurisdiction: str, source_kind: SourceKind) -> SourceDocument:
        return _load_excel_source(source_url, jurisdiction, source_kind)


def scrape_source_to_local(
    source_url: str,
    jurisdiction: str,
    source_kind: SourceKind,
    cache_dir: str | Path,
) -> Path:
    cache_path = Path(cache_dir)
    cache_path.mkdir(parents=True, exist_ok=True)
    discovery = discover_source_candidates(_fetch_html(source_url), source_url)
    resolved_url, html, resolved_discovery = _probe_content_url(source_url, discovery)
    source_slug = _slugify(f"{jurisdiction}-{source_kind.value}")
    html_path = cache_path / f"{source_slug}.html"
    html_path.write_text(html, encoding="utf-8")
    discovery_path = cache_path / f"{source_slug}.discovery.json"
    discovery_path.write_text(json.dumps(resolved_discovery, indent=2), encoding="utf-8")
    document = load_source_document_from_local(html_path, jurisdiction, source_kind)
    document_path = cache_path / f"{source_slug}.source.json"
    document_path.write_text(document.model_dump_json(indent=2, exclude_none=True), encoding="utf-8")
    chunks = [chunk.model_dump() for chunk in build_source_chunks(document)]
    chunks_path = cache_path / f"{source_slug}.chunks.json"
    chunks_path.write_text(json.dumps(chunks, indent=2), encoding="utf-8")
    meta_path = cache_path / f"{source_slug}.meta.json"
    meta_path.write_text(
        json.dumps(
            {
                "source_url": source_url,
                "resolved_source_url": resolved_url,
                "jurisdiction": jurisdiction,
                "source_kind": source_kind.value,
                "html_path": str(html_path),
                "discovery_json_path": str(discovery_path),
                "source_json_path": str(document_path),
                "chunks_json_path": str(chunks_path),
            },
            indent=2,
        ),
        encoding="utf-8",
    )
    return html_path


def _probe_content_url(source_url: str, discovery: dict[str, object]) -> tuple[str, str, dict[str, object]]:
    source_host = urlparse(source_url).netloc.lower()
    same_host_candidates = [source_url]
    other_candidates: list[str] = []
    for candidate in discovery.get("candidates", []):
        if isinstance(candidate, dict):
            url = candidate.get("url")
            if isinstance(url, str):
                candidate_host = urlparse(url).netloc.lower()
                if candidate_host == source_host:
                    same_host_candidates.append(url)
                else:
                    other_candidates.append(url)
    candidates = same_host_candidates + other_candidates

    seen: set[str] = set()
    checked: list[dict[str, object]] = []
    for candidate_url in candidates:
        if candidate_url in seen:
            continue
        seen.add(candidate_url)
        try:
            html = _fetch_html(candidate_url)
            candidate_discovery = discover_source_candidates(html, candidate_url)
            soup = BeautifulSoup(html, "lxml")
            has_content = bool(soup.find_all(["h1", "h2", "h3", "h4"]) or soup.find_all("table"))
            checked.append(
                {
                    "url": candidate_url,
                    "same_host": urlparse(candidate_url).netloc.lower() == source_host,
                    "has_content": has_content,
                    "candidate_count": candidate_discovery.get("candidate_count", 0),
                    "quality": candidate_discovery.get("quality"),
                }
            )
            if has_content:
                if urlparse(candidate_url).netloc.lower() != source_host:
                    continue
                candidate_discovery["checked"] = checked
                candidate_discovery["resolved_url"] = candidate_url
                candidate_discovery["resolved"] = True
                return candidate_url, html, candidate_discovery
        except Exception as exc:
            checked.append({"url": candidate_url, "error": str(exc)})

    discovery = dict(discovery)
    discovery["checked"] = checked
    discovery["resolved"] = False
    discovery["resolved_url"] = source_url
    return source_url, _fetch_html(source_url), discovery


def discover_source_candidates(html: str, source_url: str) -> dict[str, object]:
    soup = BeautifulSoup(html, "lxml")
    source_host = urlparse(source_url).netloc.lower()
    candidates: list[dict[str, str]] = []
    seen: set[str] = set()

    def add_candidate(kind: str, url: str, label: str) -> None:
        normalized = url.strip()
        if not normalized or normalized in seen:
            return
        candidate_host = urlparse(normalized).netloc.lower()
        seen.add(normalized)
        candidates.append(
            {
                "kind": kind,
                "url": normalized,
                "label": label,
                "host": candidate_host,
                "same_host": candidate_host == source_host,
            }
        )

    for link in soup.find_all("a", href=True):
        href = link["href"]
        text = _clean_text(link.get_text(" ", strip=True))
        if not text and not href:
            continue
        if href.startswith("javascript:") or href.startswith("#"):
            continue
        add_candidate("anchor", urljoin(source_url, href), text or href)

    for link in soup.find_all("link", href=True):
        href = link["href"]
        rel = " ".join(link.get("rel", []))
        if href.startswith("javascript:") or href.startswith("#"):
            continue
        add_candidate("link", urljoin(source_url, href), rel or href)

    for script in soup.find_all("script", src=True):
        src = script["src"]
        if src.startswith("javascript:") or src.startswith("#"):
            continue
        add_candidate("script", urljoin(source_url, src), "script-src")

    # Mine inline script and text for route-like or query-like URL hints.
    raw_text = soup.get_text(" ", strip=True)
    for match in re.findall(r"https?://[^\s\"'<>]+", html):
        add_candidate("raw-url", match, "raw html url")
    for match in re.findall(r"(?:/[A-Za-z0-9._~!$&'()*+,;=:@%/-]+\\?(?:[^\\s\"'<>]+))", html):
        add_candidate("raw-path", urljoin(source_url, match), "raw html path")
    for key in ["nodeId", "publicationId", "productId", "jobId", "docId", "sectionId"]:
        for match in re.findall(rf"{key}[=:\"]+([A-Za-z0-9_-]+)", html, flags=re.I):
            add_candidate("id-hint", f"{source_url}#{key}={match}", f"{key}={match}")
    if raw_text:
        for match in re.findall(r"(?:Chapter|Article|Section)\\s+[A-Za-z0-9._-]+", raw_text, flags=re.I):
            add_candidate("section-hint", f"{source_url}#{_slugify(match)}", match)

    quality = "candidate-rich" if candidates else "no-candidates"
    return {
        "source_url": source_url,
        "source_host": source_host,
        "candidate_count": len(candidates),
        "quality": quality,
        "candidates": candidates[:200],
    }


def load_source_document_from_local(html_path: str | Path, jurisdiction: str, source_kind: SourceKind) -> SourceDocument:
    html = Path(html_path).read_text(encoding="utf-8")
    soup = BeautifulSoup(html, "lxml")
    source_url = str(Path(html_path).resolve())
    document_title = _clean_text(soup.title.get_text(" ", strip=True)) if soup.title else None
    sections = _extract_sections(soup, source_url)
    tables = _extract_tables(soup, source_url)
    definitions = _extract_definitions(soup, source_url)
    content_quality = _classify_content_quality(source_kind, None, sections, tables, definitions)
    return SourceDocument(
        jurisdiction=jurisdiction,
        source_kind=source_kind,
        source_url=source_url,
        document_title=document_title or f"Zoning ordinance for {jurisdiction}",
        sections=sections,
        tables=tables,
        definitions=definitions,
        metadata={
            "parser": "local-html",
            "resolved_source_url": str(Path(html_path).resolve()),
            "content_quality": content_quality,
            "source_note": "Parsed from a locally cached HTML source file.",
            "local_html_path": str(Path(html_path)),
            "section_count": len(sections),
            "table_count": len(tables),
            "definition_count": len(definitions),
        },
    )


def _load_excel_source(source_path: str | Path, jurisdiction: str, source_kind: SourceKind) -> SourceDocument:
    workbook_path = Path(source_path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Excel source not found: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Excel source has no rows: {workbook_path}")

    header_row_idx = None
    header_row: list[str] = []
    for idx, row in enumerate(rows[:10]):
        normalized = [str(value).strip() if value is not None else "" for value in row]
        lowered = {value.lower() for value in normalized if value}
        if {"url", "nodeid", "title"} & lowered and ({"content"} & lowered or {"body", "text", "html"} & lowered):
            header_row_idx = idx
            header_row = normalized
            break
    if header_row_idx is None:
        raise ValueError(f"Excel source does not contain a recognizable header row: {workbook_path}")

    column_map = {name.lower(): idx for idx, name in enumerate(header_row) if name}
    url_idx = column_map.get("url")
    node_id_idx = column_map.get("nodeid")
    title_idx = column_map.get("title")
    subtitle_idx = column_map.get("subtitle")
    content_idx = column_map.get("content") or column_map.get("body") or column_map.get("text") or column_map.get("html")
    if content_idx is None:
        raise ValueError(
            f"Excel source is missing a recognizable content column: {workbook_path}. "
            f"Found columns: {header_row}"
        )

    sections: list[SourceSection] = []
    table_rows: list[list[str]] = []
    for row_idx, row in enumerate(rows[header_row_idx + 1 :], start=header_row_idx + 2):
        content = _clean_text(str(row[content_idx])) if content_idx < len(row) and row[content_idx] is not None else ""
        if not content:
            continue
        title = _clean_text(str(row[title_idx])) if title_idx is not None and title_idx < len(row) and row[title_idx] is not None else ""
        subtitle = _clean_text(str(row[subtitle_idx])) if subtitle_idx is not None and subtitle_idx < len(row) and row[subtitle_idx] is not None else ""
        node_id = _clean_text(str(row[node_id_idx])) if node_id_idx is not None and node_id_idx < len(row) and row[node_id_idx] is not None else ""
        url = _clean_text(str(row[url_idx])) if url_idx is not None and url_idx < len(row) and row[url_idx] is not None else str(workbook_path)
        section_title = title or subtitle or node_id or f"Row {row_idx}"
        section_id = _slugify(node_id or section_title, fallback=f"row-{row_idx}")
        text = "\n".join(part for part in [title, subtitle, content] if part)
        sections.append(
            SourceSection(
                section_id=section_id,
                title=section_title,
                text=text,
                path=[section_title],
                cross_references=_find_cross_references(content, url),
            )
        )
        table_rows.append(
            [
                url,
                node_id,
                title,
                subtitle,
                content[:5000],
            ]
        )

    document_title = workbook_path.stem.replace("_", " ")
    source_table = SourceTable(
        table_id="export-rows",
        title="Exported ordinance rows",
        headers=["Url", "NodeId", "Title", "Subtitle", "Content"],
        rows=table_rows,
        footnotes=[],
        source_section_id=None,
    )
    content_quality = _classify_content_quality(source_kind, None, sections, [source_table], [])
    return SourceDocument(
        jurisdiction=jurisdiction,
        source_kind=source_kind,
        source_url=str(workbook_path.resolve()),
        document_title=document_title,
        sections=sections,
        tables=[source_table],
        definitions=[],
        metadata={
            "parser": "excel-export",
            "resolved_source_url": str(workbook_path.resolve()),
            "source_note": "Parsed from a Municode XLSX export workbook.",
            "local_excel_path": str(workbook_path),
            "sheet_name": workbook.sheetnames[0],
            "row_count": len(rows) - 1,
            "section_count": len(sections),
            "table_count": 1,
            "definition_count": 0,
            "content_quality": content_quality,
        },
    )


def load_source_document_from_cache(cache_dir: str | Path, jurisdiction: str, source_kind: SourceKind) -> SourceDocument:
    cache_path = Path(cache_dir)
    source_json = next(cache_path.glob("*.source.json"), None)
    if source_json is not None and source_json.exists():
        document = SourceDocument.model_validate_json(source_json.read_text(encoding="utf-8"))
        discovery_json = next(cache_path.glob("*.discovery.json"), None)
        if discovery_json is not None and discovery_json.exists():
            try:
                discovery = json.loads(discovery_json.read_text(encoding="utf-8"))
                if isinstance(document.metadata, dict):
                    document.metadata.setdefault("discovery_quality", discovery.get("quality"))
                    document.metadata.setdefault("discovery_candidate_count", discovery.get("candidate_count"))
                    document.metadata.setdefault("discovery_candidates", discovery.get("candidates", [])[:20])
                    document.metadata.setdefault("resolved_source_url", discovery.get("resolved_url"))
            except Exception:
                pass
        return document

    html_path = next(cache_path.glob("*.html"), None)
    if html_path is not None and html_path.exists():
        return load_source_document_from_local(html_path, jurisdiction, source_kind)

    raise FileNotFoundError(f"No cached source artifacts found in {cache_path}")


def load_source_chunks_from_cache(cache_dir: str | Path) -> list[SourceChunk]:
    cache_path = Path(cache_dir)
    chunks_json = next(cache_path.glob("*.chunks.json"), None)
    if chunks_json is not None and chunks_json.exists():
        raw = json.loads(chunks_json.read_text(encoding="utf-8"))
        return [SourceChunk.model_validate(item) for item in raw]
    return []


def build_source_chunks(document: SourceDocument, max_chunk_chars: int = 9000) -> list[SourceChunk]:
    chunks: list[SourceChunk] = []
    for section in document.sections:
        chunks.extend(_chunk_section(section, max_chunk_chars=max_chunk_chars))
    for table in document.tables:
        chunks.append(
            SourceChunk(
                chunk_id=table.table_id,
                title=table.title or table.table_id,
                kind="table",
                text="\n".join([" | ".join(row) for row in ([table.headers] if table.headers else []) + table.rows]),
                source_ids=[table.source_section_id] if table.source_section_id else [],
                metadata={"footnotes": table.footnotes},
            )
        )
    if not chunks:
        chunks.append(
            SourceChunk(
                chunk_id="source-shell",
                title=document.document_title or document.jurisdiction,
                kind="shell",
                text="No content sections or tables were extracted from the source page.",
                source_ids=[],
                metadata={"content_quality": document.metadata.get("content_quality")},
            )
        )
    return chunks


def _load_html_source(source_url: str, jurisdiction: str, source_kind: SourceKind) -> SourceDocument:
    parsed_url = urlparse(source_url)
    source_params = parse_qs(parsed_url.query)
    html = _fetch_html(source_url)
    discovery = discover_source_candidates(html, source_url)
    soup = BeautifulSoup(html, "lxml")
    document_title = _clean_text(soup.title.get_text(" ", strip=True)) if soup.title else None
    has_section_tree = bool(soup.find_all(["h1", "h2", "h3", "h4"]) or soup.find_all("table"))
    node_id = source_params.get("nodeId", [None])[0]

    sections = _extract_sections(soup, source_url)
    tables = _extract_tables(soup, source_url)
    definitions = _extract_definitions(soup, source_url)
    content_quality = _classify_content_quality(source_kind, node_id, sections, tables, definitions)

    if not document_title:
        document_title = f"Zoning ordinance for {jurisdiction}"

    return SourceDocument(
        jurisdiction=jurisdiction,
        source_kind=source_kind,
        source_url=source_url,
        document_title=document_title,
        sections=sections,
        tables=tables,
        definitions=definitions,
        metadata={
            "parser": "municode-html",
            "source_host": parsed_url.netloc,
            "source_path": parsed_url.path,
            "source_fragment": parsed_url.fragment or None,
            "source_query_keys": sorted(source_params.keys()),
            "node_id": node_id,
            "has_section_tree": has_section_tree,
            "section_count": len(sections),
            "table_count": len(tables),
            "definition_count": len(definitions),
            "content_quality": content_quality,
            "discovery_quality": discovery["quality"],
            "discovery_candidate_count": discovery["candidate_count"],
            "discovery_candidates": discovery["candidates"][:20],
            "source_note": (
                "Parsed directly from a Municode content page."
                if node_id and (sections or tables)
                else (
                    "Parsed directly from HTML. Some Municode publication pages expose only the shell page "
                    "and require a specific nodeId URL for full ordinance content."
                    if not sections and not tables
                    else "Parsed from HTML content available at the supplied URL."
                )
            ),
        },
    )


def _fetch_html(url: str) -> str:
    with httpx.Client(follow_redirects=True, timeout=60.0, headers={"User-Agent": "Mozilla/5.0"}) as client:
        response = client.get(url)
        response.raise_for_status()
        return response.text


def _extract_sections(soup: BeautifulSoup, source_url: str) -> list[SourceSection]:
    sections: list[SourceSection] = []
    headings = soup.find_all(["h1", "h2", "h3", "h4"])
    for idx, heading in enumerate(headings, start=1):
        title = _clean_text(heading.get_text(" ", strip=True))
        if not title:
            continue
        text_parts: list[str] = []
        for sibling in heading.find_next_siblings():
            if sibling.name in {"h1", "h2", "h3", "h4"}:
                break
            text = _clean_text(sibling.get_text(" ", strip=True))
            if text:
                text_parts.append(text)
        text = "\n".join(text_parts).strip()
        if not text:
            continue
        section_id = _slugify(title, fallback=f"section-{idx}")
        sections.append(
            SourceSection(
                section_id=section_id,
                title=title,
                text=text,
                path=[title],
                cross_references=_find_cross_references(text, source_url),
            )
        )
    return sections


def _extract_tables(soup: BeautifulSoup, source_url: str) -> list[SourceTable]:
    parsed_tables: list[SourceTable] = []
    for idx, table in enumerate(soup.find_all("table"), start=1):
        parsed = _parse_html_table(table)
        if not parsed.headers and not parsed.rows:
            continue
        title = parsed.title or f"Table {idx}"
        parsed_tables.append(
            SourceTable(
                table_id=_slugify(title, fallback=f"table-{idx}"),
                title=title,
                headers=parsed.headers,
                rows=parsed.rows,
                footnotes=parsed.footnotes,
                source_section_id=_nearest_heading_id(table),
            )
        )
    return parsed_tables


def _extract_definitions(soup: BeautifulSoup, source_url: str) -> list[DefinedTerm]:
    definitions: list[DefinedTerm] = []
    definition_lists = soup.find_all("dl")
    for dl in definition_lists:
        terms = dl.find_all("dt")
        defs = dl.find_all("dd")
        for dt, dd in zip(terms, defs, strict=False):
            term = _clean_text(dt.get_text(" ", strip=True))
            definition = _clean_text(dd.get_text(" ", strip=True))
            if not term or not definition:
                continue
            definitions.append(
                DefinedTerm(
                    term=term,
                    definition=definition,
                    citations=[
                        Citation(
                            section_title=term,
                            quote=definition[:500],
                            source_url=source_url,
                            confidence=0.85,
                        )
                    ],
                )
            )
    return definitions


def _chunk_section(section: SourceSection, max_chunk_chars: int) -> list[SourceChunk]:
    if len(section.text) <= max_chunk_chars:
        return [
            SourceChunk(
                chunk_id=section.section_id,
                title=section.title,
                kind="section",
                text=section.text,
                source_ids=[section.section_id],
                metadata={"cross_references": section.cross_references, "path": section.path},
            )
        ]

    chunks: list[SourceChunk] = []
    parts = _split_text(section.text, max_chunk_chars)
    for idx, part in enumerate(parts, start=1):
        chunks.append(
            SourceChunk(
                chunk_id=f"{section.section_id}-{idx}",
                title=f"{section.title} (part {idx})",
                kind="section-part",
                text=part,
                source_ids=[section.section_id],
                metadata={"cross_references": section.cross_references, "path": section.path},
            )
        )
    return chunks


def _parse_html_table(table_tag) -> _ParsedTable:
    rows = table_tag.find_all("tr")
    parsed_rows: list[list[str]] = []
    headers: list[str] = []
    for row_idx, tr in enumerate(rows):
        cells = [ _clean_text(cell.get_text(" ", strip=True)) for cell in tr.find_all(["th", "td"]) ]
        if not cells:
            continue
        if row_idx == 0 and tr.find_all("th"):
            headers = cells
        else:
            parsed_rows.append(cells)
    title = _nearest_caption(table_tag)
    footnotes = _extract_footnotes(table_tag)
    return _ParsedTable(title=title, headers=headers, rows=parsed_rows, footnotes=footnotes)


def _nearest_caption(table_tag) -> str | None:
    caption = table_tag.find("caption")
    if caption:
        text = _clean_text(caption.get_text(" ", strip=True))
        if text:
            return text
    previous = table_tag.find_previous(["h1", "h2", "h3", "h4", "h5"])
    if previous:
        text = _clean_text(previous.get_text(" ", strip=True))
        if text:
            return text
    return None


def _nearest_heading_id(table_tag) -> str | None:
    previous = table_tag.find_previous(["h1", "h2", "h3", "h4"])
    if previous:
        return _slugify(previous.get_text(" ", strip=True))
    return None


def _extract_footnotes(table_tag) -> list[str]:
    notes = []
    for sup in table_tag.find_all("sup"):
        text = _clean_text(sup.get_text(" ", strip=True))
        if text:
            notes.append(text)
    return notes


def _find_cross_references(text: str, source_url: str) -> list[str]:
    refs = set()
    for match in re.findall(r"(?:Section|Sec\.|Chapter)\s+([A-Za-z0-9.\-]+)", text, flags=re.I):
        refs.add(urljoin(source_url, f"#{match}"))
    return sorted(refs)


def _clean_text(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def _classify_content_quality(
    source_kind: SourceKind,
    node_id: str | None,
    sections: list[SourceSection],
    tables: list[SourceTable],
    definitions: list[DefinedTerm],
) -> str:
    if sections or tables or definitions:
        if source_kind == SourceKind.MUNICODE and node_id:
            return "content-page"
        return "html-content"
    if source_kind == SourceKind.MUNICODE and not node_id:
        return "shell-page"
    return "empty-or-unsupported"


def _slugify(text: str, fallback: str | None = None) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")
    return slug or (fallback or "section")


def _split_text(text: str, max_chunk_chars: int) -> list[str]:
    paragraphs = [p.strip() for p in text.split("\n") if p.strip()]
    if not paragraphs:
        return [text]
    chunks: list[str] = []
    current = ""
    for para in paragraphs:
        if not current:
            current = para
        elif len(current) + len(para) + 1 <= max_chunk_chars:
            current += "\n" + para
        else:
            chunks.append(current)
            current = para
    if current:
        chunks.append(current)
    return chunks
