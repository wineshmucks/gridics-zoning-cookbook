from __future__ import annotations

from dataclasses import dataclass
import logging
from pathlib import Path
import re
from typing import Any, Iterable

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from zoning_agno.db.models import MuniNodeORM, SourceDocumentORM
from zoning_agno.schemas import MuniNode, SourceDocument, SourceKind

logger = logging.getLogger(__name__)

EXPECTED_COLUMNS = {"url", "node_id", "title", "subtitle", "content"}
COLUMN_ALIASES = {
    "nodeid": "node_id",
    "node_id": "node_id",
    "node id": "node_id",
    "url": "url",
    "title": "title",
    "subtitle": "subtitle",
    "content": "content",
}


@dataclass(slots=True)
class IngestionStats:
    source_document_id: int
    jurisdiction: str
    sheet_name: str
    row_count: int
    populated_node_ids: int
    populated_titles: int
    populated_content_rows: int
    detected_columns: list[str]


def normalize_column_name(value: Any) -> str:
    text = "" if value is None else str(value).strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text).strip("_")
    return COLUMN_ALIASES.get(text, text)


def _trim_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _score_sheet(rows: list[dict[str, Any]]) -> tuple[int, int]:
    if not rows:
        return (0, 0)
    header_matches = sum(1 for key in rows[0] if key in EXPECTED_COLUMNS)
    content_rows = sum(1 for row in rows if any(_trim_text(row.get(key)) for key in EXPECTED_COLUMNS))
    return (header_matches, content_rows)


def _score_header_row(values: tuple[Any, ...]) -> tuple[int, int]:
    normalized = [normalize_column_name(value) for value in values if normalize_column_name(value)]
    matches = sum(1 for value in normalized if value in EXPECTED_COLUMNS)
    populated = sum(1 for value in normalized if value)
    return (matches, populated)


def detect_primary_sheet(workbook_path: str | Path) -> str:
    sheets = _read_workbook_sheets(workbook_path)
    ranked = sorted(
        ((name, *_score_sheet(rows)) for name, rows in sheets.items()),
        key=lambda item: (item[1], item[2], item[0].lower()),
        reverse=True,
    )
    if not ranked:
        raise ValueError(f"No readable worksheets found in {workbook_path}")
    return ranked[0][0]


def _read_workbook_sheets(workbook_path: str | Path) -> dict[str, list[dict[str, Any]]]:
    workbook = load_workbook(filename=workbook_path, read_only=True, data_only=True)
    sheets: dict[str, list[dict[str, Any]]] = {}
    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            continue
        header_row_index, header_values = max(
            (
                (idx, row)
                for idx, row in enumerate(rows[:25], start=1)
            ),
            key=lambda item: _score_header_row(item[1]),
        )
        header = [normalize_column_name(value) or f"column_{idx + 1}" for idx, value in enumerate(header_values)]
        parsed_rows: list[dict[str, Any]] = []
        for workbook_row_number, raw_row in enumerate(rows[header_row_index:], start=header_row_index + 1):
            row = {header[idx]: raw_row[idx] if idx < len(raw_row) else None for idx in range(len(header))}
            if any(value not in (None, "") for value in row.values()):
                row["_source_row_number"] = workbook_row_number
                parsed_rows.append(row)
        sheets[worksheet.title] = parsed_rows
    workbook.close()
    return sheets


def iter_workbook_rows(workbook_path: str | Path, sheet_name: str | None = None) -> tuple[str, list[dict[str, Any]]]:
    sheets = _read_workbook_sheets(workbook_path)
    if not sheets:
        raise ValueError(f"No readable worksheets found in {workbook_path}")
    selected_sheet = sheet_name or detect_primary_sheet(workbook_path)
    if selected_sheet not in sheets:
        raise ValueError(f"Worksheet '{selected_sheet}' not found in {workbook_path}")
    return selected_sheet, sheets[selected_sheet]


def build_muni_nodes(rows: Iterable[dict[str, Any]]) -> list[MuniNode]:
    nodes: list[MuniNode] = []
    for index, row in enumerate(rows, start=2):
        normalized_row = {
            (key if str(key).startswith("_") else normalize_column_name(key)): value for key, value in row.items()
        }
        node = MuniNode(
            row_number=int(normalized_row.get("_source_row_number") or index),
            node_id=_trim_text(normalized_row.get("node_id")),
            url=_trim_text(normalized_row.get("url")),
            title=_trim_text(normalized_row.get("title")),
            subtitle=_trim_text(normalized_row.get("subtitle")),
            content=_trim_text(normalized_row.get("content")),
            raw_payload_json={key: value for key, value in normalized_row.items()},
        )
        nodes.append(node)
    return nodes


def ingest_workbook(
    session: Session,
    workbook_path: str | Path,
    jurisdiction: str,
    source_url: str | None = None,
    source_type: SourceKind = SourceKind.MUNICODE,
    sheet_name: str | None = None,
) -> IngestionStats:
    workbook_path = Path(workbook_path)
    selected_sheet, rows = iter_workbook_rows(workbook_path, sheet_name=sheet_name)
    nodes = build_muni_nodes(rows)

    source_document = SourceDocumentORM(
        jurisdiction=jurisdiction,
        source_type=source_type.value,
        source_file_name=workbook_path.name,
        source_url=source_url or str(workbook_path.resolve()),
    )
    session.add(source_document)
    session.flush()

    for node in nodes:
        session.add(
            MuniNodeORM(
                source_document_id=source_document.id,
                row_number=node.row_number,
                node_id=node.node_id,
                url=node.url,
                title=node.title,
                subtitle=node.subtitle,
                content=node.content,
                raw_payload_json=node.raw_payload_json,
            )
        )

    session.commit()
    logger.info(
        "Loaded workbook '%s' into source_document_id=%s with %s rows from sheet '%s'",
        workbook_path,
        source_document.id,
        len(nodes),
        selected_sheet,
    )
    return IngestionStats(
        source_document_id=source_document.id,
        jurisdiction=jurisdiction,
        sheet_name=selected_sheet,
        row_count=len(nodes),
        populated_node_ids=sum(1 for node in nodes if node.node_id),
        populated_titles=sum(1 for node in nodes if node.title),
        populated_content_rows=sum(1 for node in nodes if node.content),
        detected_columns=sorted({key for row in rows for key in row}),
    )


def source_document_from_orm(record: SourceDocumentORM) -> SourceDocument:
    return SourceDocument(
        id=record.id,
        jurisdiction=record.jurisdiction,
        source_type=record.source_type,
        source_file_name=record.source_file_name,
        source_url=record.source_url or "",
        imported_at=record.imported_at,
    )
