from __future__ import annotations

import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook


def _write_template(path: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Workbook Summary"
    general = workbook.create_sheet("Zones - General")
    general.append(["Field Name", "DB Field Name", "Data Type", "AO", "GC"])
    general.append(["Height", "PrincipalMaxHeightFt", "float", 45, 40])
    uses = workbook.create_sheet("Zones - Uses")
    uses.append(["Uses Label Id", "Use Name", "Use Value", "AO", "GC"])
    uses.append([373379, "Accessory Dwelling Unit", "Use Allowance", "P", "P"])
    workbook.save(path)


def test_blank_template_dynamic_cells_script_blanks_only_district_values(tmp_path: Path) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "blank.xlsx"
    _write_template(template_path)

    subprocess.run(
        [
            sys.executable,
            "scripts/blank_template_dynamic_cells.py",
            str(template_path),
            str(output_path),
        ],
        check=True,
        cwd=Path(__file__).resolve().parents[1],
    )

    workbook = load_workbook(output_path)
    general = workbook["Zones - General"]
    assert general.cell(row=1, column=1).value == "Field Name"
    assert general.cell(row=2, column=1).value == "Height"
    assert general.cell(row=2, column=2).value == "PrincipalMaxHeightFt"
    assert general.cell(row=2, column=4).value is None
    assert general.cell(row=2, column=5).value is None

    uses = workbook["Zones - Uses"]
    assert uses.cell(row=2, column=1).value == 373379
    assert uses.cell(row=2, column=2).value == "Accessory Dwelling Unit"
    assert uses.cell(row=2, column=4).value is None
    assert uses.cell(row=2, column=5).value is None
