from __future__ import annotations

from collections import Counter
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from zoning_agno.services.template_reader import build_template_row_key
from zoning_agno.services.workbook_compare import TARGET_SHEETS, normalize_cell_value


@dataclass(slots=True)
class SheetMismatch:
    sheet_name: str
    row_key: str
    row_label: str
    district_code: str
    kind: str
    generated_value: Any
    reference_value: Any
    row_index: int
    column_index: int


@dataclass(slots=True)
class SheetMismatchSummary:
    sheet_name: str
    mismatch_count: int
    by_kind: dict[str, int]
    by_row_key: dict[str, int]
    mismatches: list[SheetMismatch]


def collect_sheet_mismatches(
    generated_path: str | Path,
    reference_path: str | Path,
    *,
    sheet_names: list[str] | None = None,
) -> list[SheetMismatchSummary]:
    generated = load_workbook(Path(generated_path), data_only=True)
    reference = load_workbook(Path(reference_path), data_only=True)
    summaries: list[SheetMismatchSummary] = []

    for sheet_name in sheet_names or TARGET_SHEETS:
        if sheet_name not in generated.sheetnames or sheet_name not in reference.sheetnames:
            continue
        summaries.append(_collect_sheet_mismatch_summary(generated[sheet_name], reference[sheet_name]))
    return summaries


def summary_to_json_ready(summary: SheetMismatchSummary) -> dict[str, Any]:
    return {
        "sheet_name": summary.sheet_name,
        "mismatch_count": summary.mismatch_count,
        "by_kind": summary.by_kind,
        "by_row_key": summary.by_row_key,
        "mismatches": [asdict(item) for item in summary.mismatches],
    }


def _collect_sheet_mismatch_summary(generated_sheet, reference_sheet) -> SheetMismatchSummary:
    mismatches: list[SheetMismatch] = []
    by_kind: Counter[str] = Counter()
    by_row_key: Counter[str] = Counter()
    max_row = max(generated_sheet.max_row, reference_sheet.max_row)
    max_col = max(generated_sheet.max_column, reference_sheet.max_column)

    for row_index in range(2, max_row + 1):
        col_a = reference_sheet.cell(row_index, 1).value
        col_b = reference_sheet.cell(row_index, 2).value
        col_c = reference_sheet.cell(row_index, 3).value
        row_key = build_template_row_key(reference_sheet.title, col_a, col_b, col_c)
        row_label = _stringify(col_b or col_a or row_key)

        for column_index in range(4, max_col + 1):
            district_code = _stringify(reference_sheet.cell(1, column_index).value)
            if not district_code:
                continue
            generated_value = normalize_cell_value(generated_sheet.cell(row_index, column_index).value)
            reference_value = normalize_cell_value(reference_sheet.cell(row_index, column_index).value)
            kind = _classify_difference(generated_value, reference_value)
            if kind is None:
                continue
            mismatch = SheetMismatch(
                sheet_name=reference_sheet.title,
                row_key=row_key,
                row_label=row_label,
                district_code=district_code,
                kind=kind,
                generated_value=generated_value,
                reference_value=reference_value,
                row_index=row_index,
                column_index=column_index,
            )
            mismatches.append(mismatch)
            by_kind[kind] += 1
            by_row_key[row_key] += 1

    return SheetMismatchSummary(
        sheet_name=reference_sheet.title,
        mismatch_count=len(mismatches),
        by_kind=dict(by_kind),
        by_row_key=dict(by_row_key.most_common()),
        mismatches=mismatches,
    )


def _classify_difference(generated_value: Any, reference_value: Any) -> str | None:
    if generated_value in (None, "") and reference_value in (None, ""):
        return None
    if generated_value == reference_value:
        return None
    if generated_value in (None, ""):
        return "missing_from_generated"
    if reference_value in (None, ""):
        return "extra_in_generated"
    return "value_mismatch"


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()
