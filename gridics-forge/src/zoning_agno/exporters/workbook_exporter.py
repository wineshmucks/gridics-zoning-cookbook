from __future__ import annotations

from dataclasses import asdict, is_dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from zoning_agno.models.schemas import ExtractionBatch, ReviewFlag, SourceDocument
from zoning_agno.schemas import DistrictRegistryBundle, PivotTableBundle, SheetPopulationBundle, TemplateSheet
from zoning_agno.services.template_reader import build_template_row_key, read_template_sheet


class WorkbookExporter:
    def __init__(self, template_path: str | Path) -> None:
        self.template_path = Path(template_path)
        self.last_export_stats: dict[str, dict[str, int]] = {}

    def export(
        self,
        extraction: ExtractionBatch,
        out_path: str | Path,
        *,
        district_registry: DistrictRegistryBundle | None = None,
        general_pivot: PivotTableBundle | None = None,
        general_population: SheetPopulationBundle | None = None,
        use_population: SheetPopulationBundle | None = None,
        parking_population: SheetPopulationBundle | None = None,
        use_capacity_population: SheetPopulationBundle | None = None,
        bonus_population: SheetPopulationBundle | None = None,
        use_pivot: PivotTableBundle | None = None,
        parking_pivot: PivotTableBundle | None = None,
        overlay_pivot: PivotTableBundle | None = None,
    ) -> Path:
        out = Path(out_path)
        out.parent.mkdir(parents=True, exist_ok=True)
        wb = load_workbook(self.template_path)

        self._write_source_summary_sheet(wb, extraction)
        self._write_evidence_sheet(wb, extraction)
        self.last_export_stats = self._write_extraction_sheets(
            wb,
            extraction,
            district_registry=district_registry,
            general_pivot=general_pivot,
            general_population=general_population,
            use_population=use_population,
            parking_population=parking_population,
            use_capacity_population=use_capacity_population,
            bonus_population=bonus_population,
            use_pivot=use_pivot,
            parking_pivot=parking_pivot,
            overlay_pivot=overlay_pivot,
        )
        self._write_summary_sheet(wb, extraction)
        self._write_traceability_sheet(
            wb,
            [bundle for bundle in [general_pivot, use_pivot, parking_pivot, overlay_pivot] if bundle],
            [bundle for bundle in [general_population, use_population, parking_population, use_capacity_population, bonus_population] if bundle],
        )
        self._write_review_queue(wb, extraction.review_flags, sheet_name="Human Review Queue")
        self._write_review_queue(wb, extraction.review_flags, sheet_name="Review Queue")

        wb.save(out)
        return out

    def _write_summary_sheet(self, wb, extraction: ExtractionBatch) -> None:
        ws = self._get_or_create_sheet(wb, "Workbook Summary")
        ws.delete_rows(1, ws.max_row)
        ws.append(["Workbook Summary"])
        ws.append(["This workbook was generated from a zoning source and normalized into canonical extraction tables."])
        ws.append([])
        ws.append(["Table", "Records", "Notes"])
        ws.append(["Districts", len(extraction.districts), "Canonical district list."])
        ws.append(["General Standards", len(extraction.general_standards), "Dimensional and related standards."])
        ws.append(["Use Rules", len(extraction.use_rules), "Use allowance and restriction rules."])
        ws.append(["Parking Rules", len(extraction.parking_rules), "Parking formulas and normalized values."])
        ws.append(["Bonus Rules", len(extraction.bonus_rules), "Bonus and incentive rules."])
        ws.append(["Review Flags", len(extraction.review_flags), "Items requiring human review."])
        ws.append(["Evidence Rows", self._count_citations(extraction), "Total citations captured across all records."])
        for sheet_name, stats in self.last_export_stats.items():
            ws.append([f"{sheet_name} - Written Cells", stats.get("written_cells", 0), "Cells written from extracted populations."])
            ws.append([f"{sheet_name} - Skipped Conflicts", stats.get("skipped_conflicts", 0), "Extracted cells skipped because the template already had a different value."])
            ws.append([f"{sheet_name} - Missing Targets", stats.get("missing_targets", 0), "Extracted cells that could not map to a template row/column."])

    def _write_source_summary_sheet(self, wb, extraction: ExtractionBatch) -> None:
        ws = self._get_or_create_sheet(wb, "Source Summary")
        ws.delete_rows(1, ws.max_row)
        ws.append(["Source Summary"])
        ws.append(["The workflow source document is preserved as structured data before extraction."])
        ws.append([])
        ws.append(["Section", "Description"])
        ws.append(["Jurisdiction", extraction.metadata.get("jurisdiction", "See pipeline input")])
        ws.append(["Source URL", extraction.metadata.get("source_url", "See pipeline input")])
        ws.append(["Source Kind", extraction.metadata.get("source_kind", "See pipeline input")])
        ws.append(["Document Title", extraction.metadata.get("document_title", "If discovered during intake")])
        ws.append(["Effective Date", extraction.metadata.get("effective_date", "If discovered during intake")])
        ws.append(["Parser", extraction.metadata.get("parser", "unknown")])
        ws.append(["Content Quality", extraction.metadata.get("content_quality", "unknown")])
        ws.append(["Section Count", extraction.metadata.get("section_count", 0)])
        ws.append(["Table Count", extraction.metadata.get("table_count", 0)])
        ws.append(["Definition Count", extraction.metadata.get("definition_count", 0)])
        ws.append(["Source Note", extraction.metadata.get("source_note", "")])

    def _write_evidence_sheet(self, wb, extraction: ExtractionBatch) -> None:
        ws = self._get_or_create_sheet(wb, "Evidence")
        ws.delete_rows(1, ws.max_row)
        ws.append(["Evidence Inventory"])
        ws.append(["Each row below is a source citation attached to a generated record."])
        ws.append([])
        ws.append(["Record Type", "Record Key", "Section Id", "Section Title", "Quote", "Source URL", "Confidence"])
        for record_type, record_key, citations in self._iter_citations(extraction):
            for citation in citations:
                ws.append(
                    [
                        record_type,
                        record_key,
                        citation.section_id,
                        citation.section_title,
                        citation.quote,
                        citation.source_url,
                        citation.confidence,
                    ]
                )

    def _write_extraction_sheets(
        self,
        wb,
        extraction: ExtractionBatch,
        *,
        district_registry: DistrictRegistryBundle | None,
        general_pivot: PivotTableBundle | None,
        general_population: SheetPopulationBundle | None,
        use_population: SheetPopulationBundle | None,
        parking_population: SheetPopulationBundle | None,
        use_capacity_population: SheetPopulationBundle | None,
        bonus_population: SheetPopulationBundle | None,
        use_pivot: PivotTableBundle | None,
        parking_pivot: PivotTableBundle | None,
        overlay_pivot: PivotTableBundle | None,
    ) -> dict[str, dict[str, int]]:
        stats: dict[str, dict[str, int]] = {}
        if general_population is not None:
            template_sheet = read_template_sheet(self.template_path, "Zones - General")
            stats["Zones - General"] = self.write_population_to_sheet(wb["Zones - General"], template_sheet, general_population)
        elif general_pivot is not None:
            self._write_pivot_sheet(wb, general_pivot)
        else:
            self._write_general_matrix_sheet(wb, extraction)
        if use_population is not None:
            template_sheet = read_template_sheet(self.template_path, "Zones - Uses")
            stats["Zones - Uses"] = self.write_population_to_sheet(wb["Zones - Uses"], template_sheet, use_population)
        elif use_pivot is not None:
            self._write_pivot_sheet(wb, use_pivot)
        else:
            self._write_rows_sheet(
                wb,
                "Zones - Uses",
                extraction.use_rules,
                ["district_code", "use_label_id", "use_name", "use_value", "family", "conditions", "exceptions", "citations"],
                ["District", "Use Label Id", "Use Name", "Use Value", "Family", "Conditions", "Exceptions", "Citations"],
        )
        if parking_population is not None:
            template_sheet = read_template_sheet(self.template_path, "Zones - Parking")
            stats["Zones - Parking"] = self.write_population_to_sheet(wb["Zones - Parking"], template_sheet, parking_population)
        elif parking_pivot is not None:
            self._write_pivot_sheet(wb, parking_pivot)
        else:
            self._write_rows_sheet(
                wb,
                "Zones - Parking",
                extraction.parking_rules,
                ["district_code", "field_name", "db_field_name", "formula_text", "normalized_value", "family", "citations"],
                ["District", "Field", "DB Field", "Formula", "Normalized Value", "Family", "Citations"],
            )
        if overlay_pivot is not None:
            self._write_pivot_sheet(wb, overlay_pivot)
        if use_capacity_population is not None:
            template_sheet = read_template_sheet(self.template_path, "Zones - Use Capacity")
            stats["Zones - Use Capacity"] = self.write_population_to_sheet(wb["Zones - Use Capacity"], template_sheet, use_capacity_population)
        else:
            self._write_rows_sheet(
                wb,
                "Zones - Use Capacity",
                extraction.general_standards,
                ["district_code", "field_name", "db_field_name", "data_type", "value", "unit", "family", "citations"],
                ["District", "Field", "DB Field", "Type", "Value", "Unit", "Family", "Citations"],
            )
        if bonus_population is not None:
            template_sheet = read_template_sheet(self.template_path, "Zones - Bonus")
            stats["Zones - Bonus"] = self.write_population_to_sheet(wb["Zones - Bonus"], template_sheet, bonus_population)
        else:
            self._write_rows_sheet(
                wb,
                "Zones - Bonus",
                extraction.bonus_rules,
                ["district_code", "field_name", "db_field_name", "value", "family", "citations"],
                ["District", "Field", "DB Field", "Value", "Family", "Citations"],
            )
        return stats

    def _write_general_matrix_sheet(self, wb, extraction: ExtractionBatch) -> None:
        ws = self._get_or_create_sheet(wb, "Zones - General")
        # Preserve the draft-style matrix layout, only filling in values when we have them.
        district_headers = [cell.value for cell in ws[1][3:] if cell.value]
        field_to_row = {}
        for row_idx in range(2, ws.max_row + 1):
            field_name = ws.cell(row=row_idx, column=1).value
            db_field_name = ws.cell(row=row_idx, column=2).value
            if field_name and db_field_name:
                field_to_row[str(db_field_name)] = row_idx

        district_map = {district: idx + 4 for idx, district in enumerate(district_headers)}
        for record in extraction.general_standards:
            row_idx = field_to_row.get(record.db_field_name)
            if row_idx is None:
                continue
            district_col = district_map.get(record.district_code)
            if district_col is None:
                continue
            ws.cell(row=row_idx, column=district_col).value = self._stringify(record.value)

        # Add citations in metadata cells to the far right if present in the template footer area.
        # The draft matrix doesn't have a dedicated citation column, so we keep the sheet shape intact.

    def write_population_to_sheet(self, ws, template_sheet: TemplateSheet, population_bundle: SheetPopulationBundle) -> dict[str, int]:
        district_col_indexes = self.find_district_column_indexes(template_sheet)
        row_lookup = self.build_template_row_index_lookup(template_sheet)
        written_cells = 0
        skipped_conflicts = 0
        missing_targets = 0
        for population in population_bundle.populated_cells:
            row_index = row_lookup.get(population.row_key)
            column_index = district_col_indexes.get(population.district_code)
            if row_index is None or column_index is None:
                missing_targets += 1
                continue
            cell = ws.cell(row=row_index, column=column_index)
            existing_value = cell.value
            if existing_value not in (None, "") and existing_value != population.value:
                skipped_conflicts += 1
                continue
            cell.value = population.value
            written_cells += 1
        return {
            "written_cells": written_cells,
            "skipped_conflicts": skipped_conflicts,
            "missing_targets": missing_targets,
        }

    def find_district_column_indexes(self, template_sheet: TemplateSheet) -> dict[str, int]:
        return {header: index for index, header in enumerate(template_sheet.header.district_headers, start=4)}

    def build_template_row_index_lookup(self, template_sheet: TemplateSheet) -> dict[str, int]:
        return {row.key: row.row_index for row in template_sheet.rows}

    def _write_pivot_sheet(self, wb, bundle: PivotTableBundle) -> None:
        ws = self._get_or_create_sheet(wb, bundle.sheet_name)
        ws.delete_rows(1, ws.max_row)
        ws.append([bundle.sheet_name])
        ws.append(["District-pivoted matrix generated from normalized atomic zoning facts."])
        ws.append([])
        ws.append(["Row Key", "Row Label", *bundle.district_codes])
        for row in bundle.rows:
            ws.append([row.row_key, row.row_label, *[self._stringify(row.district_values.get(code, "")) for code in bundle.district_codes]])

    def _write_traceability_sheet(
        self,
        wb,
        pivot_bundles: list[PivotTableBundle],
        population_bundles: list[SheetPopulationBundle],
    ) -> None:
        ws = self._get_or_create_sheet(wb, "Traceability")
        ws.delete_rows(1, ws.max_row)
        ws.append(
            [
                "output_sheet",
                "row_key",
                "district_code",
                "value",
                "source_node_id",
                "source_title",
                "source_quote",
                "confidence",
            ]
        )
        for bundle in pivot_bundles:
            for row in bundle.rows:
                for district_code, value in row.district_values.items():
                    if value in (None, ""):
                        continue
                    citations = row.citations_by_district.get(district_code) or []
                    if not citations:
                        ws.append([bundle.sheet_name, row.row_key, district_code, self._stringify(value), None, None, None, None])
                        continue
                    for citation in citations:
                        ws.append(
                            [
                                bundle.sheet_name,
                                row.row_key,
                                district_code,
                                self._stringify(value),
                                getattr(citation, "node_id", None),
                                getattr(citation, "title", None) or getattr(citation, "section_title", None),
                                getattr(citation, "quote", None),
                                getattr(citation, "confidence", None),
                            ]
                        )
        for bundle in population_bundles:
            for cell in bundle.populated_cells:
                if cell.value in (None, ""):
                    continue
                if not cell.citations:
                    ws.append([cell.sheet_name, cell.row_key, cell.district_code, self._stringify(cell.value), None, None, None, cell.confidence])
                    continue
                for citation in cell.citations:
                    ws.append(
                        [
                            cell.sheet_name,
                            cell.row_key,
                            cell.district_code,
                            self._stringify(cell.value),
                            getattr(citation, "node_id", None),
                            getattr(citation, "title", None) or getattr(citation, "section_title", None),
                            getattr(citation, "quote", None),
                            getattr(citation, "confidence", None),
                        ]
                    )

    def _write_review_queue(self, wb, review_flags: list[ReviewFlag], *, sheet_name: str) -> None:
        ws = self._get_or_create_sheet(wb, sheet_name)
        ws.delete_rows(1, ws.max_row)
        ws.append(["Severity", "Sheet", "Field", "Issue Type", "Reason", "Suggested Action", "District", "Family"])
        for flag in review_flags:
            ws.append(
                [
                    flag.severity.value,
                    flag.sheet_name,
                    flag.field_key,
                    flag.issue_type,
                    flag.reason,
                    flag.suggested_action,
                    flag.district_code,
                    getattr(flag.family, "value", None) if flag.family is not None else None,
                ]
            )

    def _write_rows_sheet(
        self,
        wb,
        sheet_name: str,
        records: list[Any],
        field_names: list[str],
        headers: list[str],
    ) -> None:
        ws = self._get_or_create_sheet(wb, sheet_name)
        ws.delete_rows(1, ws.max_row)
        ws.append([sheet_name])
        ws.append(["This sheet is populated by the workflow from canonical extraction records."])
        ws.append([])
        ws.append(headers)
        for record in records:
            row = [self._stringify(self._extract_field(record, field_name)) for field_name in field_names]
            ws.append(row)

    def _get_or_create_sheet(self, wb, sheet_name: str):
        if sheet_name in wb.sheetnames:
            return wb[sheet_name]
        return wb.create_sheet(sheet_name)

    def _get_value(self, record: Any, field_name: str) -> Any:
        if isinstance(record, dict):
            return record.get(field_name)
        return getattr(record, field_name, None)

    def _extract_field(self, record: Any, field_name: str) -> Any:
        if field_name == "citations":
            return self._format_citations(self._get_value(record, field_name))
        if field_name == "conditions":
            return self._format_conditions(self._get_value(record, field_name))
        if field_name == "exceptions":
            return self._format_list(self._get_value(record, field_name))
        return self._get_value(record, field_name)

    def _format_list(self, value: Any) -> str:
        if not value:
            return ""
        if isinstance(value, list):
            return "\n".join(str(v) for v in value if v is not None)
        return str(value)

    def _format_citations(self, citations: Any) -> str:
        if not citations:
            return ""
        parts = []
        for citation in citations:
            if isinstance(citation, dict):
                section_id = citation.get("section_id") or ""
                section_title = citation.get("section_title") or ""
                quote = citation.get("quote") or ""
            else:
                section_id = getattr(citation, "section_id", "")
                section_title = getattr(citation, "section_title", "")
                quote = getattr(citation, "quote", "")
            parts.append(" | ".join(p for p in [section_id, section_title, quote] if p))
        return "\n".join(parts)

    def _format_conditions(self, conditions: Any) -> str:
        if not conditions:
            return ""
        parts = []
        for condition in conditions:
            if isinstance(condition, dict):
                expression = condition.get("expression")
                notes = condition.get("notes")
            else:
                expression = getattr(condition, "expression", None)
                notes = getattr(condition, "notes", None)
            if expression and notes:
                parts.append(f"{expression} ({notes})")
            elif expression:
                parts.append(str(expression))
            elif notes:
                parts.append(str(notes))
        return "\n".join(parts)

    def _count_citations(self, extraction: ExtractionBatch) -> int:
        return sum(1 for _ in self._iter_citations(extraction))

    def _iter_citations(self, extraction: ExtractionBatch):
        for item in extraction.districts:
            yield "districts", item.code, item.citations
        for item in extraction.general_standards:
            yield "general_standards", f"{item.district_code}:{item.db_field_name}", item.citations
        for item in extraction.use_rules:
            yield "use_rules", f"{item.district_code}:{item.use_name}", item.citations
        for item in extraction.parking_rules:
            yield "parking_rules", f"{item.district_code}:{item.db_field_name}", item.citations
        for item in extraction.bonus_rules:
            yield "bonus_rules", f"{item.district_code}:{item.db_field_name}", item.citations

    def _stringify(self, value: Any) -> Any:
        if value is None:
            return None
        if isinstance(value, (str, int, float, bool)):
            return value
        if is_dataclass(value):
            return asdict(value)
        if hasattr(value, "model_dump"):
            return value.model_dump()
        if isinstance(value, list):
            return ", ".join(self._stringify(v) for v in value if v is not None)
        return str(value)
