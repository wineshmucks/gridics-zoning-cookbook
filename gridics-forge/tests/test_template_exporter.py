from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from zoning_agno.exporters.workbook_exporter import WorkbookExporter
from zoning_agno.models.schemas import ExtractionBatch
from zoning_agno.schemas import CellPopulation, SheetPopulationBundle


def _write_template(path: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Workbook Summary"
    workbook.create_sheet("Source Summary")
    workbook.create_sheet("Evidence")
    general = workbook.create_sheet("Zones - General")
    general.append(["Field Name", "DB Field Name", "Data Type", "AO", "GC"])
    general.append(["Principal Building Max. Height", "PrincipalMaxHeightFt", "float", 45, None])
    general.append(["Minimum Lot Area", "MinLotArea", "float", None, None])
    workbook.create_sheet("Zones - Uses")
    parking = workbook.create_sheet("Zones - Parking")
    parking.append(["Field Name", "DB Field Name", "Data Type", "AO", "GC"])
    parking.append(["Space per Office Area", "Office", "float", None, None])
    use_capacity = workbook.create_sheet("Zones - Use Capacity")
    use_capacity.append(["Field Name", "DB Field Name", "Data Type", "AO", "GC"])
    use_capacity.append(["Max Dwelling Units", "MaxDwellingUnits", "float", None, None])
    bonus = workbook.create_sheet("Zones - Bonus")
    bonus.append(["Field Name", "DB Field Name", "Data Type", "AO", "GC"])
    bonus.append(["Density Bonus", "DensityBonus", "float", None, None])
    workbook.create_sheet("Review Queue")
    workbook.save(path)


def test_exporter_preserves_template_structure_and_populates_general_cells(tmp_path: Path) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "out.xlsx"
    _write_template(template_path)

    exporter = WorkbookExporter(template_path)
    population = SheetPopulationBundle(
        sheet_name="Zones - General",
        row_keys=["MinLotArea", "PrincipalMaxHeightFt"],
        district_codes=["AO", "GC"],
        populated_cells=[
            CellPopulation(
                sheet_name="Zones - General",
                row_key="MinLotArea",
                district_code="GC",
                value=6000,
            )
        ],
    )
    exporter.export(ExtractionBatch(), output_path, general_population=population)

    workbook = load_workbook(output_path)
    general = workbook["Zones - General"]
    assert [general.cell(row=1, column=idx).value for idx in range(1, 6)] == ["Field Name", "DB Field Name", "Data Type", "AO", "GC"]
    assert general.cell(row=2, column=1).value == "Principal Building Max. Height"
    assert general.cell(row=2, column=4).value == 45
    assert general.cell(row=3, column=1).value == "Minimum Lot Area"
    assert general.cell(row=3, column=5).value == 6000
    assert general.cell(row=3, column=2).value == "MinLotArea"


def test_exporter_adds_traceability_for_populated_template_cells(tmp_path: Path) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "out.xlsx"
    _write_template(template_path)

    exporter = WorkbookExporter(template_path)
    population = SheetPopulationBundle(
        sheet_name="Zones - General",
        row_keys=["MinLotArea"],
        district_codes=["AO", "GC"],
        populated_cells=[
            CellPopulation(
                sheet_name="Zones - General",
                row_key="MinLotArea",
                district_code="GC",
                value=6000,
            )
        ],
    )
    exporter.export(ExtractionBatch(), output_path, general_population=population)

    workbook = load_workbook(output_path)
    trace = workbook["Traceability"]
    assert trace.cell(row=2, column=1).value == "Zones - General"
    assert trace.cell(row=2, column=2).value == "MinLotArea"
    assert trace.cell(row=2, column=3).value == "GC"
    assert trace.cell(row=2, column=4).value == 6000


def test_exporter_preserves_template_parking_sheet_with_empty_population(tmp_path: Path) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "out.xlsx"
    _write_template(template_path)

    exporter = WorkbookExporter(template_path)
    parking_population = SheetPopulationBundle(
        sheet_name="Zones - Parking",
        row_keys=["Office"],
        district_codes=["AO", "GC"],
        populated_cells=[],
    )
    exporter.export(ExtractionBatch(), output_path, parking_population=parking_population)

    workbook = load_workbook(output_path)
    parking = workbook["Zones - Parking"]
    assert [parking.cell(row=1, column=idx).value for idx in range(1, 6)] == ["Field Name", "DB Field Name", "Data Type", "AO", "GC"]
    assert parking.cell(row=2, column=1).value == "Space per Office Area"
    assert parking.cell(row=2, column=2).value == "Office"


def test_exporter_preserves_template_use_capacity_and_bonus_sheets(tmp_path: Path) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "out.xlsx"
    _write_template(template_path)

    exporter = WorkbookExporter(template_path)
    use_capacity_population = SheetPopulationBundle(
        sheet_name="Zones - Use Capacity",
        row_keys=["MaxDwellingUnits"],
        district_codes=["AO", "GC"],
        populated_cells=[],
    )
    bonus_population = SheetPopulationBundle(
        sheet_name="Zones - Bonus",
        row_keys=["DensityBonus"],
        district_codes=["AO", "GC"],
        populated_cells=[],
    )
    exporter.export(
        ExtractionBatch(),
        output_path,
        use_capacity_population=use_capacity_population,
        bonus_population=bonus_population,
    )

    workbook = load_workbook(output_path)
    use_capacity = workbook["Zones - Use Capacity"]
    assert [use_capacity.cell(row=1, column=idx).value for idx in range(1, 6)] == ["Field Name", "DB Field Name", "Data Type", "AO", "GC"]
    assert use_capacity.cell(row=2, column=2).value == "MaxDwellingUnits"
    bonus = workbook["Zones - Bonus"]
    assert [bonus.cell(row=1, column=idx).value for idx in range(1, 6)] == ["Field Name", "DB Field Name", "Data Type", "AO", "GC"]
    assert bonus.cell(row=2, column=2).value == "DensityBonus"


def test_exporter_does_not_overwrite_existing_template_value(tmp_path: Path) -> None:
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "out.xlsx"
    _write_template(template_path)

    exporter = WorkbookExporter(template_path)
    population = SheetPopulationBundle(
        sheet_name="Zones - General",
        row_keys=["PrincipalMaxHeightFt"],
        district_codes=["AO", "GC"],
        populated_cells=[
            CellPopulation(
                sheet_name="Zones - General",
                row_key="PrincipalMaxHeightFt",
                district_code="AO",
                value=10,
            )
        ],
    )
    exporter.export(ExtractionBatch(), output_path, general_population=population)

    workbook = load_workbook(output_path)
    general = workbook["Zones - General"]
    assert general.cell(row=2, column=4).value == 45
    summary = workbook["Workbook Summary"]
    rows = [
        (summary.cell(row=row, column=1).value, summary.cell(row=row, column=2).value)
        for row in range(1, summary.max_row + 1)
    ]
    assert ("Zones - General - Skipped Conflicts", 1) in rows
