from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


TARGET_SHEETS = [
    "Zones - General",
    "Zones - Uses",
    "Zones - Parking",
    "Zones - Use Capacity",
    "Zones - Bonus",
]


@dataclass(slots=True)
class CellDiff:
    row: int
    column: int
    generated_value: Any
    reference_value: Any
    kind: str


@dataclass(slots=True)
class SheetComparison:
    sheet_name: str
    rows: int
    columns: int
    total_cells: int
    exact_nonblank_matches: int
    both_blank: int
    generated_filled_cells: int
    reference_filled_cells: int
    missing_from_generated: int
    extra_in_generated: int
    value_mismatches: int
    sample_diffs: list[CellDiff]

    @property
    def populated_cell_match_rate(self) -> float:
        compared = self.exact_nonblank_matches + self.missing_from_generated + self.extra_in_generated + self.value_mismatches
        if compared == 0:
            return 1.0
        return self.exact_nonblank_matches / compared


def compare_workbooks(
    generated_path: str | Path,
    reference_path: str | Path,
    *,
    sheet_names: list[str] | None = None,
    sample_limit: int = 10,
) -> list[SheetComparison]:
    generated = load_workbook(Path(generated_path), data_only=True)
    reference = load_workbook(Path(reference_path), data_only=True)
    comparisons: list[SheetComparison] = []
    names = sheet_names or TARGET_SHEETS

    for sheet_name in names:
        if sheet_name not in generated.sheetnames or sheet_name not in reference.sheetnames:
            continue
        comparisons.append(compare_sheet(generated[sheet_name], reference[sheet_name], sample_limit=sample_limit))

    return comparisons


def compare_sheet(generated_sheet, reference_sheet, *, sample_limit: int = 10) -> SheetComparison:
    max_row = max(generated_sheet.max_row, reference_sheet.max_row)
    max_col = max(generated_sheet.max_column, reference_sheet.max_column)
    exact_nonblank_matches = 0
    both_blank = 0
    generated_filled_cells = 0
    reference_filled_cells = 0
    missing_from_generated = 0
    extra_in_generated = 0
    value_mismatches = 0
    sample_diffs: list[CellDiff] = []

    for row_index in range(1, max_row + 1):
        for column_index in range(1, max_col + 1):
            generated_value = normalize_cell_value(generated_sheet.cell(row_index, column_index).value)
            reference_value = normalize_cell_value(reference_sheet.cell(row_index, column_index).value)

            if generated_value not in (None, ""):
                generated_filled_cells += 1
            if reference_value not in (None, ""):
                reference_filled_cells += 1

            if generated_value in (None, "") and reference_value in (None, ""):
                both_blank += 1
                continue

            if generated_value == reference_value:
                exact_nonblank_matches += 1
                continue

            if generated_value in (None, "") and reference_value not in (None, ""):
                missing_from_generated += 1
                append_diff(
                    sample_diffs,
                    sample_limit,
                    row_index,
                    column_index,
                    generated_value,
                    reference_value,
                    "missing_from_generated",
                )
                continue

            if generated_value not in (None, "") and reference_value in (None, ""):
                extra_in_generated += 1
                append_diff(
                    sample_diffs,
                    sample_limit,
                    row_index,
                    column_index,
                    generated_value,
                    reference_value,
                    "extra_in_generated",
                )
                continue

            value_mismatches += 1
            append_diff(
                sample_diffs,
                sample_limit,
                row_index,
                column_index,
                generated_value,
                reference_value,
                "value_mismatch",
            )

    return SheetComparison(
        sheet_name=generated_sheet.title,
        rows=max_row,
        columns=max_col,
        total_cells=max_row * max_col,
        exact_nonblank_matches=exact_nonblank_matches,
        both_blank=both_blank,
        generated_filled_cells=generated_filled_cells,
        reference_filled_cells=reference_filled_cells,
        missing_from_generated=missing_from_generated,
        extra_in_generated=extra_in_generated,
        value_mismatches=value_mismatches,
        sample_diffs=sample_diffs,
    )


def normalize_cell_value(value: Any) -> Any:
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    return value


def append_diff(
    sample_diffs: list[CellDiff],
    sample_limit: int,
    row: int,
    column: int,
    generated_value: Any,
    reference_value: Any,
    kind: str,
) -> None:
    if len(sample_diffs) >= sample_limit:
        return
    sample_diffs.append(
        CellDiff(
            row=row,
            column=column,
            generated_value=generated_value,
            reference_value=reference_value,
            kind=kind,
        )
    )
