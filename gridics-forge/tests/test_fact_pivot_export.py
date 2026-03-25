from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from zoning_agno.db.base import Base
from zoning_agno.ingest.municode import ingest_workbook
from zoning_agno.models.schemas import ExtractionBatch
from zoning_agno.schemas import SourceKind
from zoning_agno.services.district_registry import build_district_registry
from zoning_agno.services.fact_extractor import extract_dimensional_facts
from zoning_agno.services.normalization_service import normalize_source_document
from zoning_agno.services.pivot_builder import build_general_pivot_rows
from zoning_agno.exporters.workbook_exporter import WorkbookExporter


def _write_source_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append(["Url", "NodeId", "Title", "Subtitle", "Content"])
    worksheet.append(["https://example.com/1", "GC1", "GC - General Commercial District", "", "Maximum height in GC district is 50 feet."])
    worksheet.append(["https://example.com/2", "RS61", "RS-6 Single-Family District", "", "Minimum lot area in RS-6 district is 6000 square feet."])
    workbook.save(path)


def _write_template(path: Path) -> None:
    workbook = Workbook()
    workbook.active.title = "Workbook Summary"
    for sheet_name in ["Source Summary", "Evidence", "Zones - General", "Zones - Uses", "Review Queue"]:
        workbook.create_sheet(sheet_name)
    workbook.save(path)


def test_dimensional_fact_extraction_and_general_pivot(tmp_path: Path) -> None:
    workbook_path = tmp_path / "source.xlsx"
    _write_source_workbook(workbook_path)
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    Base.metadata.create_all(engine)
    with Session(engine) as session:
        ingest_stats = ingest_workbook(
            session=session,
            workbook_path=workbook_path,
            jurisdiction="Test City",
            source_url="https://example.com/source",
            source_type=SourceKind.MUNICODE,
        )
        normalize_source_document(session, ingest_stats.source_document_id)
        registry = build_district_registry(session, ingest_stats.source_document_id)
        facts = extract_dimensional_facts(session, ingest_stats.source_document_id, registry)

    assert any(fact.district_code == "GC" and fact.field_name == "PrincipalMaxHeight" and fact.value_numeric == 50 for fact in facts.facts)
    assert any(fact.district_code == "RS-6" and fact.field_name == "MinLotArea" and fact.value_numeric == 6000 for fact in facts.facts)

    pivot = build_general_pivot_rows(registry, facts)
    assert pivot.sheet_name == "Zones - General"
    height_row = next(row for row in pivot.rows if row.row_key == "PrincipalMaxHeight")
    assert height_row.district_values["GC"] == 50
    area_row = next(row for row in pivot.rows if row.row_key == "MinLotArea")
    assert area_row.district_values["RS-6"] == 6000


def test_exporter_writes_pivoted_general_sheet_and_traceability(tmp_path: Path) -> None:
    source_path = tmp_path / "source.xlsx"
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "out.xlsx"
    _write_source_workbook(source_path)
    _write_template(template_path)
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    Base.metadata.create_all(engine)
    with Session(engine) as session:
        ingest_stats = ingest_workbook(
            session=session,
            workbook_path=source_path,
            jurisdiction="Test City",
            source_url="https://example.com/source",
            source_type=SourceKind.MUNICODE,
        )
        normalize_source_document(session, ingest_stats.source_document_id)
        registry = build_district_registry(session, ingest_stats.source_document_id)
        facts = extract_dimensional_facts(session, ingest_stats.source_document_id, registry)
    pivot = build_general_pivot_rows(registry, facts)

    exporter = WorkbookExporter(template_path)
    exporter.export(ExtractionBatch(), output_path, district_registry=registry, general_pivot=pivot)

    workbook = load_workbook(output_path)
    general = workbook["Zones - General"]
    traceability = workbook["Traceability"]
    assert general.cell(row=4, column=1).value == "Row Key"
    headers = [general.cell(row=4, column=column).value for column in range(3, 5)]
    assert headers == ["GC", "RS-6"]
    assert any(cell.value == "MinLotArea" for cell in general["A"])
    assert traceability.max_row >= 3
    assert traceability.cell(row=2, column=1).value == "Zones - General"
