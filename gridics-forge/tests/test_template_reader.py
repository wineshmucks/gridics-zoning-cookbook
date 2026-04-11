from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from zoning_agno.services.template_reader import (
    build_template_row_key,
    extract_district_headers,
    read_target_template_sheets,
    read_template_sheet,
)


def _write_template(path: Path) -> None:
    workbook = Workbook()
    general = workbook.active
    general.title = "Zones - General"
    general.append(["Field Name", "DB Field Name", "Data Type", "AO", "GC"])
    general.append(["Principal Building Max. Height", "PrincipalMaxHeightFt", "float", None, None])
    general.append(["Minimum Lot Area", "MinLotArea", "float", None, None])
    uses = workbook.create_sheet("Zones - Uses")
    uses.append(["Uses Label Id", "Use Name", "Use Value", "AO", "GC"])
    uses.append([373358, "Accessory Commercial Unit", "Use Allowance", None, None])
    workbook.create_sheet("Zones - Parking").append(["Field Name", "DB Field Name", "Data Type", "AO", "GC"])
    workbook.create_sheet("Zones - Use Capacity").append(["Field Name", "DB Field Name", "Data Type", "AO", "GC"])
    workbook.create_sheet("Zones - Bonus").append(["Field Name", "DB Field Name", "Data Type", "AO", "GC"])
    workbook.save(path)


def test_read_template_sheet_preserves_headers_and_rows(tmp_path: Path) -> None:
    template_path = tmp_path / "template.xlsx"
    _write_template(template_path)

    sheet = read_template_sheet(template_path, "Zones - General")

    assert sheet.header.fixed_headers == ["Field Name", "DB Field Name", "Data Type"]
    assert sheet.header.district_headers == ["AO", "GC"]
    assert sheet.rows[0].row_index == 2
    assert sheet.rows[0].col_a == "Principal Building Max. Height"
    assert sheet.rows[0].col_b == "PrincipalMaxHeightFt"


def test_extract_district_headers_and_row_key_rules(tmp_path: Path) -> None:
    template_path = tmp_path / "template.xlsx"
    _write_template(template_path)
    general = read_template_sheet(template_path, "Zones - General")
    uses = read_template_sheet(template_path, "Zones - Uses")

    assert extract_district_headers(general) == ["AO", "GC"]
    assert build_template_row_key("Zones - General", "Minimum Lot Area", "MinLotArea", "float") == "MinLotArea"
    assert build_template_row_key("Zones - Uses", 373358, "Accessory Commercial Unit", "Use Allowance") == "373358|Accessory Commercial Unit|Use Allowance"
    assert uses.rows[0].key == "373358|Accessory Commercial Unit|Use Allowance"


def test_read_target_template_sheets_reads_all_target_sheets(tmp_path: Path) -> None:
    template_path = tmp_path / "template.xlsx"
    _write_template(template_path)

    sheets = read_target_template_sheets(template_path)

    assert set(sheets) == {"Zones - General", "Zones - Uses", "Zones - Parking", "Zones - Use Capacity", "Zones - Bonus"}
