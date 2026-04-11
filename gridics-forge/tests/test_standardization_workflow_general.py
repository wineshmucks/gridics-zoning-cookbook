from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from zoning_agno.config import Settings
from zoning_agno.db import create_engine_from_settings, create_session_factory, initialize_database
from zoning_agno.ingest.municode import ingest_workbook
from zoning_agno.models.schemas import PipelineInput
from zoning_agno.schemas import SourceKind
from zoning_agno.services.normalization_service import normalize_source_document
from zoning_agno.workflows import build_zoning_standardization_workflow


def _write_source_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append(["Url", "NodeId", "Title", "Subtitle", "Content"])
    worksheet.append(["https://example.com/1", "GC1", "GC - General Commercial District", "", "Maximum height in GC district is 50 feet."])
    worksheet.append(["https://example.com/2", "RS61", "RS-6 Single-Family District", "", "Minimum lot area in RS-6 district is 6000 square feet."])
    workbook.save(path)


def _write_template(path: Path) -> None:
    workbook = Workbook()
    workbook.active.title = "Workbook Summary"
    workbook.create_sheet("Source Summary")
    workbook.create_sheet("Evidence")
    general = workbook.create_sheet("Zones - General")
    general.append(["Field Name", "DB Field Name", "Data Type", "GC", "RS-6"])
    general.append(["Principal Building Max. Height", "PrincipalMaxHeightFt", "float", None, None])
    general.append(["Minimum Lot Area", "MinLotArea", "float", None, None])
    workbook.create_sheet("Zones - Uses")
    parking = workbook.create_sheet("Zones - Parking")
    parking.append(["Field Name", "DB Field Name", "Data Type", "GC", "RS-6"])
    parking.append(["Space per Office Area", "Office", "float", None, None])
    use_capacity = workbook.create_sheet("Zones - Use Capacity")
    use_capacity.append(["Field Name", "DB Field Name", "Data Type", "GC", "RS-6"])
    use_capacity.append(["Max Dwelling Units", "MaxDwellingUnits", "float", None, None])
    bonus = workbook.create_sheet("Zones - Bonus")
    bonus.append(["Field Name", "DB Field Name", "Data Type", "GC", "RS-6"])
    bonus.append(["Density Bonus", "DensityBonus", "float", None, None])
    workbook.create_sheet("Review Queue")
    workbook.save(path)


def test_standardization_workflow_exports_general_template(tmp_path: Path) -> None:
    db_path = tmp_path / "workflow.db"
    source_path = tmp_path / "source.xlsx"
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "out.xlsx"
    artifact_root = tmp_path / "artifacts"
    _write_source_workbook(source_path)
    _write_template(template_path)

    settings = Settings(
        database_url=f"sqlite+pysqlite:///{db_path}",
        workbook_template_path=template_path,
    )
    engine = create_engine_from_settings(settings)
    initialize_database(engine)
    session_factory = create_session_factory(settings)
    with session_factory() as session:
        ingest_stats = ingest_workbook(
            session=session,
            workbook_path=source_path,
            jurisdiction="Test City",
            source_url="https://example.com/source",
            source_type=SourceKind.MUNICODE,
        )
        normalize_source_document(session, ingest_stats.source_document_id)

    workflow = build_zoning_standardization_workflow(settings)
    result = workflow.run(
        input=PipelineInput(
            jurisdiction="Test City",
            source_url="database://source_document",
            workbook_template_path=str(template_path),
            source_document_id=1,
        ).model_dump(),
        session_state={
            "source_document_id": 1,
            "pipeline_input": {},
            "out_path": str(output_path),
            "artifact_root": str(artifact_root),
        },
    )

    assert result.content["workbook_output_path"] == str(output_path)
    workbook = load_workbook(output_path)
    general = workbook["Zones - General"]
    assert [general.cell(row=1, column=idx).value for idx in range(1, 6)] == ["Field Name", "DB Field Name", "Data Type", "GC", "RS-6"]
    assert general.cell(row=2, column=1).value == "Principal Building Max. Height"
    assert general.cell(row=3, column=1).value == "Minimum Lot Area"
    assert general.cell(row=3, column=5).value == 6000
    use_capacity = workbook["Zones - Use Capacity"]
    assert use_capacity.cell(row=1, column=1).value == "Field Name"
    assert use_capacity.cell(row=2, column=2).value == "MaxDwellingUnits"
    bonus = workbook["Zones - Bonus"]
    assert bonus.cell(row=1, column=1).value == "Field Name"
    assert bonus.cell(row=2, column=2).value == "DensityBonus"
