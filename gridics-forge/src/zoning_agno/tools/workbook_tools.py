from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from zoning_agno.models.schemas import CanonicalWorkbookTemplate, WorkbookSheetTemplate


def load_workbook_template(path: str | Path) -> CanonicalWorkbookTemplate:
    wb = load_workbook(path)
    sheets: list[WorkbookSheetTemplate] = []
    for name in wb.sheetnames:
        ws = wb[name]
        headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
        headers = [str(h) for h in headers if h is not None]
        static_columns = headers[:3]
        dynamic_columns = headers[3:]
        sheets.append(
            WorkbookSheetTemplate(
                name=name,
                static_columns=static_columns,
                dynamic_columns=dynamic_columns,
            )
        )
    return CanonicalWorkbookTemplate(sheets=sheets)
