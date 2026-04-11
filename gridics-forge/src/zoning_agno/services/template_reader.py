from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from zoning_agno.schemas import TemplateRow, TemplateSheet, TemplateSheetHeader

TARGET_TEMPLATE_SHEETS = [
    "Zones - General",
    "Zones - Uses",
    "Zones - Parking",
    "Zones - Use Capacity",
    "Zones - Bonus",
]


def load_template_workbook(path: str | Path):
    """Load the canonical Gridics workbook template."""
    return load_workbook(Path(path))


def read_template_sheet(path: str | Path, sheet_name: str) -> TemplateSheet:
    """Read one Gridics template sheet while preserving its fixed columns, district headers, and row order."""
    workbook = load_template_workbook(path)
    worksheet = workbook[sheet_name]
    fixed_headers = [worksheet.cell(row=1, column=index).value for index in range(1, 4)]
    district_headers = extract_district_headers_from_worksheet(worksheet)
    header = TemplateSheetHeader(
        sheet_name=sheet_name,
        fixed_headers=[_normalize_header(value) for value in fixed_headers],
        district_headers=district_headers,
    )
    rows: list[TemplateRow] = []
    for row_index in range(2, worksheet.max_row + 1):
        col_a = worksheet.cell(row=row_index, column=1).value
        col_b = worksheet.cell(row=row_index, column=2).value
        col_c = worksheet.cell(row=row_index, column=3).value
        if _is_trailing_empty_row(worksheet, row_index):
            continue
        key = build_template_row_key(sheet_name, col_a, col_b, col_c)
        rows.append(
            TemplateRow(
                sheet_name=sheet_name,
                row_index=row_index,
                key=key,
                col_a=col_a,
                col_b=col_b,
                col_c=col_c,
                raw_values={
                    worksheet.cell(row=1, column=column_index).value: worksheet.cell(row=row_index, column=column_index).value
                    for column_index in range(1, worksheet.max_column + 1)
                    if worksheet.cell(row=1, column=column_index).value is not None
                },
            )
        )
    return TemplateSheet(sheet_name=sheet_name, header=header, rows=rows)


def read_target_template_sheets(path: str | Path) -> dict[str, TemplateSheet]:
    """Read the canonical Gridics output sheets exactly as structured in the workbook template."""
    return {sheet_name: read_template_sheet(path, sheet_name) for sheet_name in TARGET_TEMPLATE_SHEETS}


def extract_district_headers(template_sheet: TemplateSheet) -> list[str]:
    """Return district headers exactly as preserved from the template header row."""
    return list(template_sheet.header.district_headers)


def build_template_row_key(sheet_name: str, col_a: Any, col_b: Any, col_c: Any) -> str:
    """Build a stable row key from preserved template columns without altering row order."""
    if sheet_name == "Zones - Uses":
        return "|".join(_stringify_cell(value) for value in [col_a, col_b, col_c])
    if col_b not in (None, ""):
        return _stringify_cell(col_b)
    return "|".join(_stringify_cell(value) for value in [col_a, col_b, col_c])


def extract_district_headers_from_worksheet(worksheet) -> list[str]:
    return [
        _stringify_cell(worksheet.cell(row=1, column=column_index).value)
        for column_index in range(4, worksheet.max_column + 1)
        if worksheet.cell(row=1, column=column_index).value not in (None, "")
    ]


def _is_trailing_empty_row(worksheet, row_index: int) -> bool:
    values = [worksheet.cell(row=row_index, column=column_index).value for column_index in range(1, worksheet.max_column + 1)]
    return not any(value not in (None, "") for value in values)


def _stringify_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_header(value: Any) -> str:
    return _stringify_cell(value)
