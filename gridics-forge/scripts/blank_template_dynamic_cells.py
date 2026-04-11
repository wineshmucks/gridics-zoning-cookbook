from __future__ import annotations

from pathlib import Path

import typer
from openpyxl import load_workbook

from zoning_agno.services.template_reader import TARGET_TEMPLATE_SHEETS


app = typer.Typer(add_completion=False, help="Create a copy of the template workbook with district-value cells blanked.")


@app.command()
def main(
    template_path: Path = typer.Argument(..., exists=True, readable=True, help="Source workbook template."),
    out_path: Path = typer.Argument(..., help="Output workbook path."),
) -> None:
    workbook = load_workbook(template_path)
    for sheet_name in TARGET_TEMPLATE_SHEETS:
        if sheet_name not in workbook.sheetnames:
            continue
        worksheet = workbook[sheet_name]
        for row_index in range(2, worksheet.max_row + 1):
            for column_index in range(4, worksheet.max_column + 1):
                worksheet.cell(row=row_index, column=column_index).value = None
    out_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(out_path)
    typer.echo(f"Blanked template written to: {out_path}")


if __name__ == "__main__":
    app()
