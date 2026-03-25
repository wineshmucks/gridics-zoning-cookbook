from __future__ import annotations

from openpyxl import Workbook

from zoning_agno.services.workbook_compare import compare_workbooks


def build_workbook(path, *, general_ao=None, general_cb=None, use_ao=None) -> None:
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)

    general = workbook.create_sheet("Zones - General")
    general.append(["Field Name", "DB Field Name", "Data Type", "AO", "CB"])
    general.append(["Height", "PrincipalMaxHeightFt", "float", general_ao, general_cb])

    uses = workbook.create_sheet("Zones - Uses")
    uses.append(["Uses Label Id", "Use Name", "Use Value", "AO", "CB"])
    uses.append([373379, "Accessory Dwelling Unit", "Use Allowance", use_ao, None])

    workbook.create_sheet("Zones - Parking")
    workbook.create_sheet("Zones - Use Capacity")
    workbook.create_sheet("Zones - Bonus")
    workbook.save(path)


def test_compare_workbooks_reports_missing_extra_and_mismatched_cells(tmp_path) -> None:
    generated = tmp_path / "generated.xlsx"
    reference = tmp_path / "reference.xlsx"
    build_workbook(generated, general_ao=35, general_cb=10, use_ao="PC")
    build_workbook(reference, general_ao=35, general_cb=None, use_ao="P")

    comparisons = compare_workbooks(generated, reference, sample_limit=5)
    by_name = {comparison.sheet_name: comparison for comparison in comparisons}

    general = by_name["Zones - General"]
    assert general.exact_nonblank_matches >= 5
    assert general.extra_in_generated == 1
    assert any(diff.kind == "extra_in_generated" for diff in general.sample_diffs)

    uses = by_name["Zones - Uses"]
    assert uses.value_mismatches == 1
    assert any(diff.kind == "value_mismatch" for diff in uses.sample_diffs)
